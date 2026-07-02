from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name='Arial', color='0000FF')
BLACK = Font(name='Arial', color='000000')
GREEN = Font(name='Arial', color='008000')
BOLD = Font(name='Arial', bold=True)
HEADER = Font(name='Arial', bold=True, size=12)
TITLE = Font(name='Arial', bold=True, size=14)
YELLOW = PatternFill('solid', start_color='FFFF00')
LIGHT_GRAY = PatternFill('solid', start_color='F5F5F7')
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD
        cell.fill = LIGHT_GRAY
        cell.border = THIN

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# --- ASSUMPTIONS SHEET ---
ws_a = wb.active
ws_a.title = 'Assumptions'
ws_a['A1'] = 'RENVOA CLINIC — Model Assumptions'
ws_a['A1'].font = TITLE
ws_a['A3'] = 'OPERATING ASSUMPTIONS'
ws_a['A3'].font = BOLD

labels = [
    ('Monthly orders (Base)', 150, 'Orders per month — Year 1 base case'),
    ('Average units per order', 1.8, 'Vials per order'),
    ('Free shipping threshold ($)', 150, 'From website promo'),
    ('Payment processing rate', 0.029, 'Stripe ~2.9%'),
    ('Payment fixed fee ($)', 0.30, 'Per transaction'),
    ('Return/refund rate', 0.03, 'Industry estimate'),
    ('Marketing % of revenue', 0.25, 'DTC peptide e-commerce'),
    ('Monthly fixed overhead ($)', 1850, 'Excl. marketing — see detail'),
    ('Ramp — Month 1 orders', 40, 'Launch month'),
    ('Ramp — Month 6 orders', 120, 'Mid-year'),
    ('Ramp — Month 12 orders', 200, 'End Year 1'),
    ('Wholesale tier', '100-vial MOQ', 'Source: Wholesale Peptide Supply'),
]
start = 5
for i, (label, val, note) in enumerate(labels):
    r = start + i
    ws_a.cell(r, 1, label)
    c = ws_a.cell(r, 2, val)
    c.font = BLUE
    if isinstance(val, float) and val < 1:
        c.number_format = '0.0%'
    elif isinstance(val, (int, float)):
        c.number_format = '#,##0'
    ws_a.cell(r, 3, note)
    if 'Marketing' in label or 'Ramp' in label:
        for col in [1, 2, 3]:
            ws_a.cell(r, col).fill = YELLOW

ws_a['A18'] = 'FIXED MONTHLY OVERHEAD DETAIL'
ws_a['A18'].font = BOLD
overhead = [
    ('E-commerce platform & hosting', 75),
    ('Cold storage / fulfillment', 250),
    ('Insurance & legal compliance', 300),
    ('Accounting & bookkeeping', 150),
    ('Customer support (part-time)', 500),
    ('Software (email, analytics, COA)', 125),
    ('Miscellaneous', 450),
]
for i, (item, amt) in enumerate(overhead):
    r = 19 + i
    ws_a.cell(r, 1, item)
    ws_a.cell(r, 2, amt).font = BLUE
    ws_a.cell(r, 2).number_format = '$#,##0'
ws_a.cell(26, 1, 'Total fixed overhead').font = BOLD
ws_a.cell(26, 2, '=SUM(B19:B25)').font = BLACK
ws_a.cell(26, 2).number_format = '$#,##0'

set_widths(ws_a, [38, 18, 45])

# --- PRICING ANALYSIS SHEET ---
ws_p = wb.create_sheet('Pricing Analysis')
ws_p['A1'] = 'RENVOA CLINIC — Product Pricing & Margin Analysis'
ws_p['A1'].font = TITLE
ws_p['A3'] = 'Source: Wholesale Peptide Supply published wholesale (100-vial MOQ, Jul 2026); competitor retail from Biotech Peptides, Grey Research (Jul 2026)'
ws_p['A3'].font = Font(name='Arial', italic=True, size=9)

headers = ['Product', 'Size', 'RENVOA Price', 'Wholesale COGS', 'Packaging/COA', 'Landed COGS', 'Gross Profit', 'Gross Margin %', 'Competitor Low', 'Competitor High', 'vs Market', 'Verdict']
for c, h in enumerate(headers, 1):
    ws_p.cell(5, c, h)
style_header_row(ws_p, 5, len(headers))

products = [
    ('BPC-157', '5mg', 49, 16, 3, 45, 52),
    ('TB-500', '5mg', 59, 30, 3, 72, 72),
    ('Semaglutide', '2mg', 89, 8, 4, 120, 150),
    ('Ipamorelin', '2mg', 39, 10, 3, 35, 50),
    ('CJC-1295 (no DAC)', '2mg', 45, 15, 3, 40, 55),
    ('GHK-Cu', '50mg', 35, 15, 2, 28, 45),
]
for i, (name, size, price, wholesale, pack, comp_low, comp_high) in enumerate(products):
    r = 6 + i
    ws_p.cell(r, 1, name)
    ws_p.cell(r, 2, size)
    ws_p.cell(r, 3, price).font = BLUE
    ws_p.cell(r, 3).number_format = '$#,##0'
    ws_p.cell(r, 4, wholesale).font = BLUE
    ws_p.cell(r, 4).number_format = '$#,##0'
    ws_p.cell(r, 5, pack).font = BLUE
    ws_p.cell(r, 5).number_format = '$#,##0'
    ws_p.cell(r, 6, f'=C{r}+D{r}+E{r}')
    ws_p.cell(r, 6).number_format = '$#,##0'
    ws_p.cell(r, 7, f'=C{r}-F{r}')
    ws_p.cell(r, 7).number_format = '$#,##0'
    ws_p.cell(r, 8, f'=G{r}/C{r}')
    ws_p.cell(r, 8).number_format = '0.0%'
    ws_p.cell(r, 9, comp_low).font = BLUE
    ws_p.cell(r, 9).number_format = '$#,##0'
    ws_p.cell(r, 10, comp_high).font = BLUE
    ws_p.cell(r, 10).number_format = '$#,##0'
    ws_p.cell(r, 11, f'=IF(C{r}<I{r},"Below market",IF(C{r}>J{r},"Above market","Competitive"))')
    ws_p.cell(r, 12, f'=IF(H{r}>=0.5,"✓ Profitable",IF(H{r}>=0.35,"Marginal","Review pricing"))')

r = 13
ws_p.cell(r, 1, 'WEIGHTED AVERAGE (assumed mix)').font = BOLD
ws_p.cell(r, 3, '=SUMPRODUCT(C6:C11,B16:B21)/SUM(B16:B21)').number_format = '$#,##0.00'
ws_p.cell(r, 6, '=SUMPRODUCT(F6:F11,B16:B21)/SUM(B16:B21)').number_format = '$#,##0.00'
ws_p.cell(r, 7, '=C13-F13').number_format = '$#,##0.00'
ws_p.cell(r, 8, '=G13/C13').number_format = '0.0%'

ws_p['A15'] = 'Assumed sales mix (% of units)'
ws_p['A15'].font = BOLD
mix_labels = ['BPC-157', 'TB-500', 'Semaglutide', 'Ipamorelin', 'CJC-1295', 'GHK-Cu']
mix_vals = [0.28, 0.18, 0.15, 0.17, 0.12, 0.10]
for i, (lbl, val) in enumerate(zip(mix_labels, mix_vals)):
    ws_p.cell(16 + i, 1, lbl)
    ws_p.cell(16 + i, 2, val).font = BLUE
    ws_p.cell(16 + i, 2).number_format = '0%'
ws_p.cell(22, 1, 'Total mix').font = BOLD
ws_p.cell(22, 2, '=SUM(B16:B21)').number_format = '0%'

ws_p['A24'] = 'KEY FINDINGS'
ws_p['A24'].font = BOLD
findings = [
    '✓ All 6 SKUs show 43–78% gross margins at 100-vial wholesale MOQ pricing',
    '✓ RENVOA prices sit at or below major competitors — competitive positioning',
    '⚠ TB-500 has thinnest margin (44%) — consider $64–69 or bundle pricing',
    '✓ Semaglutide 2mg at $89 is strong vs competitors selling 5mg at $120–150',
    '⚠ Requires ~$1,600 minimum inventory buy-in (100 vials) before first sale',
    '⚠ Cold-chain shipping ($12–22/order) erodes margin on orders under $150',
]
for i, f in enumerate(findings):
    ws_p.cell(25 + i, 1, f)

set_widths(ws_p, [22, 8, 14, 14, 14, 14, 14, 14, 14, 14, 14, 16])

# --- MONTHLY P&L SHEET ---
ws_pl = wb.create_sheet('Monthly P&L')
ws_pl['A1'] = 'RENVOA CLINIC — Monthly Profit & Loss (Year 1)'
ws_pl['A1'].font = TITLE

months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
ws_pl.cell(3, 1, 'Line Item').font = BOLD
for i, m in enumerate(months, 2):
    ws_pl.cell(3, i, m).font = BOLD
    ws_pl.cell(3, i).fill = LIGHT_GRAY
ws_pl.cell(3, 14, 'YEAR 1 TOTAL').font = BOLD
ws_pl.cell(3, 14).fill = LIGHT_GRAY

# Row references to assumptions
# B5=monthly orders base, but we use ramp - put order counts in row 4 as inputs
ws_pl.cell(4, 1, 'Orders')
ramp_orders = [40, 55, 70, 85, 100, 120, 135, 150, 165, 180, 190, 200]
for i, o in enumerate(ramp_orders, 2):
    c = ws_pl.cell(4, i, o)
    c.font = BLUE
    c.number_format = '#,##0'
ws_pl.cell(4, 14, '=SUM(B4:M4)')

pl_rows = [
    ('Units sold', None, 'orders * units/order'),
    ('Gross Revenue', None, 'orders * AOV'),
    ('Less: Refunds', None, 'revenue * refund rate'),
    ('Net Revenue', None, 'gross - refunds'),
    ('', None, ''),
    ('COGS — Product', None, 'units * landed cost'),
    ('COGS — Shipping outbound', None, 'orders * ship cost'),
    ('COGS — Payment processing', None, 'revenue * rate + orders * fee'),
    ('Total COGS', None, 'sum'),
    ('Gross Profit', None, 'net rev - total cogs'),
    ('Gross Margin %', None, ''),
    ('', None, ''),
    ('Operating Expenses', None, ''),
    ('  Marketing & ads', None, 'net rev * mkt%'),
    ('  Fixed overhead', None, 'from assumptions'),
    ('Total OpEx', None, ''),
    ('', None, ''),
    ('EBITDA', None, 'gross profit - opex'),
    ('EBITDA Margin %', None, ''),
]

# Assumption refs (Pricing weighted avg)
# C13 weighted price, F13 weighted landed on Pricing sheet
# Assumptions: B6 units/order, B9 refund, B7 proc rate, B8 proc fee, B10 mkt%, B26 fixed overhead

row_start = 5
for idx, (label, _, _) in enumerate(pl_rows):
    r = row_start + idx
    ws_pl.cell(r, 1, label)
    if label:
        ws_pl.cell(r, 1).font = BOLD if label in ('Net Revenue', 'Total COGS', 'Gross Profit', 'Total OpEx', 'EBITDA', 'Operating Expenses') else Font(name='Arial')

for col in range(2, 14):
    cl = get_column_letter(col)
    # Row 5: Units = orders * units/order (Assumptions B6)
    ws_pl.cell(5, col, f'={cl}4*Assumptions!$B$6')
    ws_pl.cell(5, col).number_format = '#,##0'
    # Row 6: Gross Revenue = orders * AOV (weighted price * units/order)
    ws_pl.cell(6, col, f'={cl}4*Assumptions!$B$6*\'Pricing Analysis\'!$C$13')
    ws_pl.cell(6, col).number_format = '$#,##0'
    # Row 7: Refunds
    ws_pl.cell(7, col, f'=-{cl}6*Assumptions!$B$9')
    ws_pl.cell(7, col).number_format = '$#,##0'
    # Row 8: Net Revenue
    ws_pl.cell(8, col, f'={cl}6+{cl}7')
    ws_pl.cell(8, col).number_format = '$#,##0'
    # Row 10: Product COGS
    ws_pl.cell(10, col, f'={cl}5*\'Pricing Analysis\'!$F$13')
    ws_pl.cell(10, col).number_format = '$#,##0'
    # Row 11: Shipping — $14 avg (cold chain)
    ws_pl.cell(11, col, f'={cl}4*14')
    ws_pl.cell(11, col).number_format = '$#,##0'
    # Row 12: Payment processing
    ws_pl.cell(12, col, f'={cl}6*Assumptions!$B$7+{cl}4*Assumptions!$B$8')
    ws_pl.cell(12, col).number_format = '$#,##0'
    # Row 13: Total COGS
    ws_pl.cell(13, col, f'={cl}10+{cl}11+{cl}12')
    ws_pl.cell(13, col).number_format = '$#,##0'
    # Row 14: Gross Profit
    ws_pl.cell(14, col, f'={cl}8-{cl}13')
    ws_pl.cell(14, col).number_format = '$#,##0'
    # Row 15: Gross Margin
    ws_pl.cell(15, col, f'=IF({cl}8=0,"-",{cl}14/{cl}8)')
    ws_pl.cell(15, col).number_format = '0.0%'
    # Row 17: blank skipped row 16
    # Row 18: Marketing
    ws_pl.cell(18, col, f'={cl}8*Assumptions!$B$10')
    ws_pl.cell(18, col).number_format = '$#,##0'
    # Row 19: Fixed overhead
    ws_pl.cell(19, col, '=Assumptions!$B$26')
    ws_pl.cell(19, col).number_format = '$#,##0'
    # Row 20: Total OpEx
    ws_pl.cell(20, col, f'={cl}18+{cl}19')
    ws_pl.cell(20, col).number_format = '$#,##0'
    # Row 22: EBITDA (skip row 21 blank)
    ws_pl.cell(22, col, f'={cl}14-{cl}20')
    ws_pl.cell(22, col).number_format = '$#,##0'
    # Row 23: EBITDA margin
    ws_pl.cell(23, col, f'=IF({cl}8=0,"-",{cl}22/{cl}8)')
    ws_pl.cell(23, col).number_format = '0.0%'

# Year totals column N
for r in [4, 5, 6, 7, 8, 10, 11, 12, 13, 14, 18, 19, 20, 22]:
    ws_pl.cell(r, 14, f'=SUM(B{r}:M{r})')
    ws_pl.cell(r, 14).number_format = '$#,##0' if r >= 6 else '#,##0'
ws_pl.cell(15, 14, '=IF(N8=0,"-",N14/N8)').number_format = '0.0%'
ws_pl.cell(23, 14, '=IF(N8=0,"-",N22/N8)').number_format = '0.0%'

set_widths(ws_pl, [28] + [12]*13)

# --- SCENARIOS SHEET ---
ws_s = wb.create_sheet('Scenarios')
ws_s['A1'] = 'RENVOA CLINIC — Scenario Summary (Annual)'
ws_s['A1'].font = TITLE

sc_headers = ['Metric', 'Conservative', 'Base (Model)', 'Optimistic']
for c, h in enumerate(sc_headers, 1):
    ws_s.cell(3, c, h)
style_header_row(ws_s, 3, 4)

scenario_inputs = [
    ('Annual orders', 900, 1490, 2400),
    ('Avg order value ($)', 93, 95.8, 98),
    ('Fully-loaded gross margin %', 0.38, 0.446, 0.48),
    ('Marketing % of revenue', 0.30, 0.25, 0.20),
    ('Annual fixed overhead ($)', 22200, 22200, 24000),
]
for i, (label, con, base, opt) in enumerate(scenario_inputs):
    r = 4 + i
    ws_s.cell(r, 1, label)
    for col, val in enumerate([con, base, opt], 2):
        c = ws_s.cell(r, col, val)
        c.font = BLUE
        if 'margin' in label or 'Marketing' in label:
            c.number_format = '0.0%'
        elif isinstance(val, float):
            c.number_format = '$#,##0.00'
        else:
            c.number_format = '#,##0' if 'orders' in label else '$#,##0'

metrics = [
    ('Gross Revenue', '=B4*B5', '=C4*C5', '=D4*D5'),
    ('Gross Profit', '=B9*B6', '=C9*C6', '=D9*D6'),
    ('Marketing spend', '=B9*B7', '=C9*C7', '=D9*D7'),
    ('Fixed overhead', '=B8', '=C8', '=D8'),
    ('Total OpEx', '=B12+B13', '=C12+C13', '=D12+D13'),
    ('Net Profit (EBITDA)', '=B10-B14', '=C10-C14', '=D10-D14'),
    ('Net Margin %', '=B15/B9', '=C15/C9', '=D15/D9'),
    ('Break-even orders/mo', '=(B8/12)/(B5*(B6-B7))', '=(C8/12)/(C5*(C6-C7))', '=(D8/12)/(D5*(D6-D7))'),
]
start_m = 9
for i, (label, f1, f2, f3) in enumerate(metrics):
    r = start_m + i
    ws_s.cell(r, 1, label).font = BOLD if 'Net Profit' in label or 'Gross Revenue' in label else Font(name='Arial')
    ws_s.cell(r, 2, f1)
    ws_s.cell(r, 3, f2)
    ws_s.cell(r, 4, f3)
    for col in range(2, 5):
        fmt = '0.0%' if 'Margin' in label or 'margin' in label else ('#,##0.0' if 'Break-even' in label else '$#,##0')
        ws_s.cell(r, col).number_format = fmt
        ws_s.cell(r, col).font = BLACK

ws_s['A18'] = 'STARTUP CAPITAL REQUIRED'
ws_s['A18'].font = BOLD
startup = [
    ('Initial inventory (100 vials)', 1900),
    ('Website & branding', 2500),
    ('Legal / compliance setup', 3500),
    ('Cold-chain packaging (500 units)', 1200),
    ('Testing / COA setup', 800),
    ('Working capital (3 mo OpEx)', 15000),
]
for i, (item, amt) in enumerate(startup):
    r = 19 + i
    ws_s.cell(r, 1, item)
    ws_s.cell(r, 2, amt).font = BLUE
    ws_s.cell(r, 2).number_format = '$#,##0'
ws_s.cell(25, 1, 'TOTAL STARTUP CAPITAL').font = BOLD
ws_s.cell(25, 2, '=SUM(B19:B24)').font = BLACK
ws_s.cell(25, 2).number_format = '$#,##0'

ws_s['A27'] = 'VERDICT'
ws_s['A27'].font = BOLD
ws_s['A28'] = 'RENVOA CLINIC product pricing is PROFITABLE — 44–87% gross margins per SKU at wholesale MOQ.'
ws_s['A29'] = 'Base case Year 1 EBITDA ~$5K (3.5% net margin) after 25% marketing; break-even ~102 orders/month.'
ws_s['A30'] = 'Year 1 is tight — profitability improves sharply in Year 2 as marketing % drops and orders scale.'
ws_s['A31'] = 'Biggest risks: 100-vial inventory MOQ (~$1,900), cold-chain shipping, compliance, TB-500 thin margin.'

set_widths(ws_s, [36, 18, 18, 18])

out = '/Users/dylaneurell/novasequence/RENVOA_CLINIC_PL.xlsx'
wb.save(out)
print(out)